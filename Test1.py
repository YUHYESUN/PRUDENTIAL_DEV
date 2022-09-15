from bs4 import BeautifulSoup
import os
import urllib.request
from requests import get
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.by import By
import math
import time

'''
(공지사항 , 퇴직연금공시 , 상품공시) 이외 메뉴 실행 
하단 주석 제외 후 실행
'''

pruMainUrl = "https://www.prudential.co.kr"

dataTempltExcel = load_workbook('dataTemplate.xlsx')#엑셀 템플릿
dataTempltExcel0908 = load_workbook('DataTemplate_0908.xlsx')#엑셀 템플릿_추가요구사항 0908 기준
# chromeDriver = webdriver.Chrome(ChromeDriverManager().install())
noMakeDirMenu = ['13339'] #폴더가 생성되지 않아도 되는 공시실 id
noMakeDirTabId = ['donation' , 'social-service' ,'variable-insurance-product-disclosure'] #폴더가 생성되지 않아도 되는 탭 id

def selectTab(menuId , mainUrl , tabIdList) : 

    url = mainUrl
    html = getPageSourceHtml(mainUrl) # html을 문자열로 가져온다.
            
    h2Nm = html.find("h2" , {"class" : "carousel__item-heading"})  #메뉴명 ex)상품공시 , 경영공시 ..
    
    #최상위폴더 생성
    if menuId not in noMakeDirMenu :
        h2NmStrip = checkExistPathOrFile("output" + "/" + h2Nm.text.strip())  #동일한 파일명 있는지 확인
        os.mkdir(h2NmStrip)

    for tabId in tabIdList : 
        apiUrl = mainUrl + "?tab=" + tabId
        html = getPageSourceHtml(apiUrl) # html을 문자열로 가져온다.
        
        tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기
        tabNm = tabInfo.find("a" , {"class" : "accordion-tabs__item-toggle"}).find("span").text.strip()

        if menuId not in noMakeDirMenu and tabId not in noMakeDirTabId :
            tabPath = checkExistPathOrFile(h2NmStrip + "/" + tabNm)  #동일한 파일명 있는지 확인
            os.mkdir(tabPath)
        
        if menuId == "13342":   #상품공시

            if tabId == "currently-selling" :   #옵션아이디가 탭별로 다름
                optionId = "ddlContract"
            elif tabId == "discontinued" :
                optionId = "ddlDiscontinuedContract"

            optionSelect = tabInfo.find("select" , id = optionId)
            options = optionSelect.findAll("option")  #옵션 목록

            #주계약 (01), 특약 (02) url 2번 돌아야함
            for option in options : 
                optionVal = option["value"]
                
                optionPath = checkExistPathOrFile(tabPath + "/" + option.text)  #동일한 파일명 있는지 확인
                os.mkdir(optionPath)
                url = apiUrl + "&ct=" + optionVal

                html = getPageSourceHtml(url)# html을 문자열로 가져온다.

                tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기

                product(tabInfo , optionPath)

        elif menuId == "13343":  #변액공시
            if tabId == 'insurance-disclosure-at-any-time' :
                  #페이지로 구성되어 있음 
                  #페이지 1씩 더하다가 체크된 페이지랑 url에 입력된 페이지랑 맞지 않으면 스탑
                insertPage = "1"
                currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text

                while insertPage == currentPage :  #없는 페이지이면 스탑
                    insertPage = str(int(insertPage) + 1)
                    variableInsuranceAccordian(tabInfo , tabPath)

                    url = apiUrl + "&variableinsurance=" + insertPage
                    html = getPageSourceHtml(url)# html을 문자열로 가져온다.
                    tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기
                    currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text
                
            elif tabId == 'operating-manual' or tabId == 'trust-terms' :
                variableInsurance(tabInfo , tabPath)

            elif tabId == 'variable-insurance-product-disclosure' : 

                 variableInsuranceBoxList(tabInfo)

            elif tabId == 'risk-indicator' :
                #페이지로 구성되어 있음 
                #페이지 1씩 더하다가 체크된 페이지랑 url에 입력된 페이지랑 맞지 않으면 스탑
                insertPage = "1"
                currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text

                while insertPage == currentPage :  #없는 페이지이면 스탑
                    insertPage = str(int(insertPage) + 1)
                    variableInsuranceRiskAccordian(tabInfo , tabPath)

                    url = apiUrl + "&riskindicator=" + insertPage
                    html = getPageSourceHtml(url)# html을 문자열로 가져온다.
                    tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기
                    currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text
        
        elif menuId == '13347': #경영공시
            if tabId == 'regular' or tabId == 'governance' :
                cmpyInformationTable(tabInfo , tabPath , tabId)
            elif tabId == 'occasional' :
                #페이지로 구성되어 있음 
                #페이지 1씩 더하다가 체크된 페이지랑 url에 입력된 페이지랑 맞지 않으면 스탑
                insertPage = "1"
                currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text

                while insertPage == currentPage :  #없는 페이지이면 스탑
                    insertPage = str(int(insertPage) + 1)
                    cmpyInformationAccordian(tabInfo , tabPath )

                    url = apiUrl + "&page=" + insertPage
                    html = getPageSourceHtml(url)# html을 문자열로 가져온다.
                    tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기
                    currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text

        elif menuId == '13348': #사회공헌공시
            if tabId == 'donation' or tabId == 'social-service' :

                socialContributionAccordian(tabInfo , tabId)

            elif tabId == 'regulations' or tabId == 'disclosure' :

                socialContribution(tabInfo , tabPath , tabId)

        elif menuId == '13339' : #사회공헌활동 연혁
            socialContributionHis(tabInfo)

    return

def variableInsuranceBoxList(tabInfo) :
    variableInsuranceBoxList = tabInfo.select('.panel__block')[0]
    boxList = variableInsuranceBoxList.select('.box-list__col')

    tabList = ["variable-product-tab2", "variable-product-tab3"] #자산구성 / 자산부채
    
    #자세히 보기 
    for box in boxList :
        routeUrl = box.find("a")["href"]

        prodAll = box.find("div" , {"class" : "box-list__box-heading --small"}).text.strip()
        prodList = prodAll.split("\n\n")

        if routeUrl.find(pruMainUrl) == -1 :
            routeUrl = pruMainUrl + routeUrl

        clickVariableInsurance(routeUrl , tabList , prodList)

    return

def clickVariableInsurance(url , tabList , prodList) :

    for tab in tabList :
        url = url + "&tab=" + tab
        chromeDriver.get(url)
        time.sleep(5)

        #변액보험은 js렌더링 후 데이터 불러옴
        tabContents = chromeDriver.find_element(By.ID, tab)
        tabTable = tabContents.find_element(By.CLASS_NAME , "panel__block")
        if tab == "variable-product-tab2" :
            sheetPath = dataTempltExcel.get_sheet_by_name("자산구성내역(변액보험)")   #엑셀 시트명
            sheetName = "자산구성내역"
            yyyy = tabTable.find_elements(By.CLASS_NAME , "panel__block")[0].find_element(By.CLASS_NAME,"lblYear").text
            mm = tabTable.find_elements(By.CLASS_NAME , "panel__block")[0].find_element(By.CLASS_NAME,"lblMonth").text 
            panelContents = tabTable.find_elements(By.CLASS_NAME , "panel__block")[1]
            productList = panelContents.find_elements(By.CLASS_NAME , "panel__block")
        elif tab == "variable-product-tab3" :
            sheetPath = dataTempltExcel.get_sheet_by_name("자산부채현황(변액보험)")   #엑셀 시트명
            sheetName = "자산부채현황"
            yyyy = tabTable.find_element(By.CLASS_NAME,"lblYear").text
            mm = tabTable.find_element(By.CLASS_NAME,"lblMonth").text 
            productList = tabTable.find_elements(By.CLASS_NAME , "panel__block")
        row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 

        for product in productList :
            typeNm = product.find_element(By.TAG_NAME,"p").text
            content = product.find_element(By.TAG_NAME,"table").get_attribute("outerHTML")

            for prodNm in prodList :

                setExcelValueVI(sheetPath , row , '공시구분' , "변액보험") #엑셀 셀 값 저장(공시구분)
                setExcelValueVI(sheetPath , row , '현황구분' , sheetName) #엑셀 셀 값 저장(현황구분)
                setExcelValueVI(sheetPath , row , '상품명' , prodNm) #엑셀 셀 값 저장(상품명)
                setExcelValueVI(sheetPath , row , '구분(펀드)' , typeNm) #엑셀 셀 값 저장(구분(펀드))
                setExcelValueVI(sheetPath , row , '기준년' , yyyy) #엑셀 셀 값 저장(기준년)
                setExcelValueVI(sheetPath , row , '기준일' , mm) #엑셀 셀 값 저장(기준월)
                setExcelValueVI(sheetPath , row , '상품코드' , str(content)) #엑셀 셀 값 저장(기준월)

                print("success : " , sheetName , "=> " , typeNm, "/" ,prodNm)

                row += 1

            dataTempltExcel.save('output/test.xlsx')  #엑셀 다른이름 저장  
           
    return

def variableInsuranceAccordian(tabInfo , tabPath) : # tapPath : 변액보험공시/변액보험수시공시
    variableInsuranceAccordian = tabInfo.select('.panel__block')[0]
    accordianList = variableInsuranceAccordian.select('.accordion')
    sheetPath = dataTempltExcel.get_sheet_by_name("변액보험수시공시")   #엑셀 시트명
    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 

    for accordian in accordianList :
        rgstP = accordian.find("p" , {"class": "accordion__cover-small"} )
        rgstDt = rgstP.text.replace("공시 일자 ","").strip()
        rgstYYYY = rgstDt[:4]
        rgstMM = rgstDt[5:7]

        yyyyPath = tabPath + "/" + rgstYYYY 
        os.makedirs(yyyyPath , exist_ok= True)

        mmPath = yyyyPath + "/" + rgstMM + "월"
        os.makedirs(mmPath , exist_ok= True)

        lastPath = mmPath + "/" + rgstDt
        os.makedirs(lastPath , exist_ok= True)

        accordianContents = accordian.find("div" , {"class": "accordion__contents"})
        mainTopic = accordian.find("a" , {"class": "accordion__pointer"} )["title"]
        sheetPath.cell(row,2).value = mainTopic    # 제목
        sheetPath.cell(row,3).value = str(accordianContents) # 내용
        sheetPath.cell(row,4).value = rgstDt.replace("-","")    # 등록일자

        fileList = accordianContents.findAll("a")
        imgList = accordianContents.findAll("img")
        filePathListStr = ""

        for file in fileList:
            if str(file).find('getattachment') == -1: # 첨부파일아니면 넘어가
                continue
            fileDownLoadUrl = file["href"].strip()        #다운로드할 파일 url

            if fileDownLoadUrl.find('www.prudential.co.kr') == -1:
                fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
            
            fileImg = file.find("img")
            fileFont = file.find("font")
            if fileImg != None : # 첨부파일명
                saveName = fileImg["alt"]
                downloadPath = lastPath + "/" + saveName      #저장 경로
            elif fileFont != None : # 펀드명
                fileNm = fileDownLoadUrl[fileDownLoadUrl.rfind('/')+1:]
                fileNm = fileNm[:fileNm.rfind('.')] # VA혼합형_2022_2.pdf
                filePath = lastPath + "/fund/" + fileFont.text
                os.makedirs(filePath , exist_ok= True)
                saveName = fileNm
                downloadPath = filePath + "/" + saveName      #저장 경로

            try:
                download(fileDownLoadUrl , downloadPath)
                filePathListStr += downloadPath + ","
                print("success : " , downloadPath)
            except urllib.error.HTTPError as e:
                print("failed:", e)

        for file in imgList:
            if str(file).find('getattachment') == -1: # 첨부파일 아니면 넘어가
                continue
            fileDownLoadUrl = file["src"].strip()        #다운로드할 파일 url

            if fileDownLoadUrl.find('www.prudential.co.kr') == -1:
                fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
            saveName = fileDownLoadUrl[fileDownLoadUrl.rfind('/')+1:]
            saveName = saveName[:saveName.rfind('.')] # 신규펀드-공시-캡처.PNG

            filePath = lastPath + "/img"
            os.makedirs(filePath , exist_ok= True)
            downloadPath = filePath + "/" + saveName      #저장 경로

            try:
                download(fileDownLoadUrl , downloadPath)
                filePathListStr += downloadPath + ","
                print("success : " , downloadPath)
            except urllib.error.HTTPError as e:
                print("failed:", e)

        sheetPath.cell(row,5).value = filePathListStr    # 첨부파일 경로명
        row += 1

    dataTempltExcel.save('output/variableAnyTime.xlsx')  #엑셀 다른이름 저장 

    return


def variableInsuranceRiskAccordian(tabInfo , tabPath) :
    variableInsuranceRiskAccordian = tabInfo.select('.panel__block')[0]
    accordianList = variableInsuranceRiskAccordian.select('.accordion')
    sheetPath = dataTempltExcel0908.get_sheet_by_name("위험지표공시")   #엑셀 시트명
    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 

    for accordian in accordianList :
        rgstP = accordian.find("p" , {"class": "accordion__cover-small"} )
        rgstDt = rgstP.text.replace("공시 일자 ","").strip()
        rgstYYYY = rgstDt[:4]
        lastPath = tabPath + "/" + rgstYYYY 
        os.makedirs(lastPath , exist_ok= True)
        
        accordianContents = accordian.find("div" , {"class": "accordion__contents"})
        mainTopic = accordian.find("a" , {"class": "accordion__pointer"} )["title"]
        setExcelValue(sheetPath , row , '제목(공시제목)' , mainTopic) #엑셀 셀 값 저장(제목)
        setExcelValue(sheetPath , row , '공시일자' , rgstDt.replace("-","").strip()) #엑셀 셀 값 저장(공시일자)
        setExcelValue(sheetPath , row , '내용(HTML)' , str(accordianContents)) #엑셀 셀 값 저장(내용)

        fileList = accordianContents.findAll("a")
        fileIndex = 1

        for file in fileList:
            if str(file).find('getattachment') == -1: # 첨부파일아니면 넘어가
                continue
            fileDownLoadUrl = file["href"].strip()        #다운로드할 파일 url

            if fileDownLoadUrl.find('www.prudential.co.kr') == -1:
                fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
            
            fileImg = file.find("img")
            if fileImg != None : # 첨부파일명
                saveName = fileImg["alt"]
                downloadPath = lastPath + "/" + saveName      #저장 경로

            try:
                download(fileDownLoadUrl , downloadPath)
                dowmloadColNm = '첨부파일' + str(fileIndex)
                setExcelValue(sheetPath , row , dowmloadColNm , downloadPath) #엑셀 셀 값 저장(첨부파일)
                print("success : " , downloadPath)
                fileIndex += 1
            except urllib.error.HTTPError as e:
                print("failed:", e)

        row += 1

    dataTempltExcel0908.save('output/com0908_2.xlsx')  #엑셀 다른이름 저장 
    return

def socialContributionAccordian(tabInfo , tabId) : #사회공헌공시 아코디언 (기부 , 봉사활동)
    socialContributionAccordian = tabInfo.select('.panel__block')[0]
    accordianList = socialContributionAccordian.select('.accordion')

    if tabId == 'donation' :

        sheetPath = dataTempltExcel.get_sheet_by_name("사회공헌기부공시")   #엑셀 시트명

    elif tabId == 'social-service' :

        sheetPath = dataTempltExcel.get_sheet_by_name("사회공헌활동공시")   #엑셀 시트명

    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 

    for accordian in accordianList :
        yyyy = accordian.find("a")['title'].replace("년" , "").strip()
    
        accordianContents = accordian.find("div" , {"class" : "accordion__contents"})
        contentsTableList = accordianContents.select(".table-holder tbody")
        if len(contentsTableList) == 1 :
            contentsTable = accordianContents.select(".table-holder tbody")[0]
        else : 
            contentsTable = accordianContents.select(".table-holder tbody")[1]
        
        contentsTables = contentsTable.findAll("tr")

        startFor = 0
        for contents in contentsTables :
            
            month = contents.find("th") #진행시기
            
            if month != None :
                monthStrip = month.text.replace("월", "").strip() #공백제거

                if monthStrip == "합계" :
                    continue

                try :
                    monthQ = math.ceil(int(monthStrip)/3)
                except ValueError : 
                    print("-------------------------------" , yyyy)
                    setExcelValue(sheetPath , row , '년도' , yyyy) #엑셀 셀 값 저장(년도)
                    setExcelValue(sheetPath , row , '장소' , contentsTd[0].text) #엑셀 셀 값 저장(장소)
                    setExcelValue(sheetPath , row , '비고' , "VALUEERROR") #엑셀 셀 값 저장(비고)

                    row += 1
                    continue

                try : 
                    rowLen = int(month['rowspan'])
                except KeyError :
                    rowLen = 1
                

                realI = startFor

                for i in range(startFor, startFor + rowLen):

                    if realI != i :
                        continue

                    contentsTd = contentsTables[i].findAll("td") #세부내용 (기부항목 , 집행액 , 집행목적 등등)
                    setExcelValue(sheetPath , row , '년도' , yyyy) #엑셀 셀 값 저장(년도)
                    setExcelValue(sheetPath , row , '분기' , str(monthQ) + "Q") #엑셀 셀 값 저장(분기)
                    setExcelValue(sheetPath , row , '월' , monthStrip) #엑셀 셀 값 저장(월)

                    if tabId == 'donation' :

                        try :
                            tdRowLen = int(contentsTd[0]['rowspan'])

                            setExcelValue(sheetPath , row , '기부항목' , contentsTd[0].text) #엑셀 셀 값 저장(기부항목)
                            setExcelValue(sheetPath , row , '기부/집행액(백만원)' , contentsTd[1].text) #엑셀 셀 값 저장(기부/집행액(백만원))
                            setExcelValue(sheetPath , row , '기부/집행목적' , contentsTd[2].text) #엑셀 셀 값 저장(기부/집행목적)
                            setExcelValue(sheetPath , row , '비고' , contentsTd[3].text) #엑셀 셀 값 저장(비고)
                            row += 1 #로우 증가

                            print("success : " , yyyy , monthStrip , contentsTd[0].text , "=> " , contentsTd[2].text)

                            for j in range(i + 1 , i + tdRowLen):

                                contentsType = contentsTables[j].findAll("td") #세부내용 (기부항목 , 집행액 , 집행목적 등등)
                                setExcelValue(sheetPath , row , '년도' , yyyy) #엑셀 셀 값 저장(년도)
                                setExcelValue(sheetPath , row , '분기' , str(monthQ) + "Q") #엑셀 셀 값 저장(분기)
                                setExcelValue(sheetPath , row , '월' , monthStrip) #엑셀 셀 값 저장(월)
                                setExcelValue(sheetPath , row , '기부항목' , contentsTd[0].text) #엑셀 셀 값 저장(기부항목)
                                setExcelValue(sheetPath , row , '기부/집행액(백만원)' , contentsType[0].text) #엑셀 셀 값 저장(기부/집행액(백만원))
                                setExcelValue(sheetPath , row , '기부/집행목적' , contentsType[1].text) #엑셀 셀 값 저장(기부/집행목적)
                                setExcelValue(sheetPath , row , '비고' , contentsType[2].text) #엑셀 셀 값 저장(비고)

                                print("success : " , yyyy , monthStrip , contentsTd[0].text , "=> " , contentsType[1].text)

                                row += 1 #로우 증가
                            realI += tdRowLen -1 
                            
                        except KeyError :

                            setExcelValue(sheetPath , row , '기부항목' , contentsTd[0].text) #엑셀 셀 값 저장(기부항목)
                            setExcelValue(sheetPath , row , '기부/집행액(백만원)' , contentsTd[1].text) #엑셀 셀 값 저장(기부/집행액(백만원))
                            setExcelValue(sheetPath , row , '기부/집행목적' , contentsTd[2].text) #엑셀 셀 값 저장(기부/집행목적)
                            setExcelValue(sheetPath , row , '비고' , contentsTd[3].text) #엑셀 셀 값 저장(비고)

                            print("success : " , yyyy , monthStrip , contentsTd[0].text , "=> " , contentsTd[2].text)
                            row += 1 #로우 증가
                    
                    elif tabId == 'social-service' :
                        contentsNm = contentsTables[i].find("th", {"style" : "text-align: left;"}) #th가 두개 이상인 경우
                        
                        if len(contentsTd) == 6 :
                            setExcelValue(sheetPath , row , '장소' , contentsNm.text) #엑셀 셀 값 저장(장소)
                            setExcelValue(sheetPath , row , '봉사활동' , contentsTd[0].text) #엑셀 셀 값 저장(봉사활동)
                            setExcelValue(sheetPath , row , '참석인원 임직원 시간' , contentsTd[1].text) #엑셀 셀 값 저장(참석인원 임직원 시간)
                            setExcelValue(sheetPath , row , '참석인원 임직원 인원' , contentsTd[2].text) #엑셀 셀 값 저장(참석인원 임직원 인원)
                            setExcelValue(sheetPath , row , '참석인원 설계사 시간' , contentsTd[3].text) #엑셀 셀 값 저장(참석인원 설계사 시간)
                            setExcelValue(sheetPath , row , '참석인원 설계사 인원' , contentsTd[4].text) #엑셀 셀 값 저장(참석인원 설계사 인원)
                            setExcelValue(sheetPath , row , '비고' , contentsTd[5].text) #엑셀 셀 값 저장(비고)
                            
                            print("success : " , yyyy , monthStrip , contentsNm.text , "=> " , contentsTd[0].text)
                        # else :
                        elif len(contentsTd) == 7 :
                            setExcelValue(sheetPath , row , '장소' , contentsTd[0].text) #엑셀 셀 값 저장(장소)
                            setExcelValue(sheetPath , row , '봉사활동' , contentsTd[1].text) #엑셀 셀 값 저장(봉사활동)
                            setExcelValue(sheetPath , row , '참석인원 임직원 시간' , contentsTd[2].text) #엑셀 셀 값 저장(참석인원 임직원 시간)
                            setExcelValue(sheetPath , row , '참석인원 임직원 인원' , contentsTd[3].text) #엑셀 셀 값 저장(참석인원 임직원 인원)
                            setExcelValue(sheetPath , row , '참석인원 설계사 시간' , contentsTd[4].text) #엑셀 셀 값 저장(참석인원 설계사 시간)
                            setExcelValue(sheetPath , row , '참석인원 설계사 인원' , contentsTd[5].text) #엑셀 셀 값 저장(참석인원 설계사 인원)
                            setExcelValue(sheetPath , row , '비고' , contentsTd[6].text) #엑셀 셀 값 저장(비고)

                            print("success : " , yyyy , monthStrip , contentsTd[0].text , "=> " , contentsTd[1].text)

                        row += 1 #로우 증가

                    realI += 1
                
                startFor += rowLen
                

    dataTempltExcel.save('output/test.xlsx')  #엑셀 다른이름 저장 

    return

def socialContributionHis(tabInfo) :

    sheetPath = dataTempltExcel.get_sheet_by_name("사회공헌소식")   #엑셀 시트명
    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 
    timelineList = tabInfo.select('.timeline__item')
    
    for timeline in timelineList:
        yyyy = timeline.find("h4").text.strip().replace("년","")

        contentsList = timeline.find("ul" , {"class" : "bullet-list"}).findAll("li")
        
        for contents in contentsList :
            setExcelValue(sheetPath , row , '제목' , contents.text.strip()) #엑셀 셀 값 저장(제목)
            setExcelValue(sheetPath , row , '등록일자' , yyyy) #엑셀 셀 값 저장(등록일자)
            print("success : " , yyyy , "=> ", contents.text.strip())
            row += 1 #로우 증가

    dataTempltExcel.save('output/test.xlsx')  #엑셀 다른이름 저장 

    return

def cmpyInformationAccordian(tabInfo , tabPath ) : #경영공시 아코디언 형식
    cmpyInformationAccordian = tabInfo.select('.panel__block')[0]
    accordianList = cmpyInformationAccordian.select('.accordion')
    sheetPath = dataTempltExcel.get_sheet_by_name("경영공시(정기,수시,결산,감사)")   #엑셀 시트명
    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 

    for accordian in accordianList :
        rgstP = accordian.find("p" , {"class": "accordion__cover-small"} )
        rgstDt = rgstP.text.strip().replace("등록일 ","")
        rgstYYYY = rgstDt[:4]
        rgstMM = rgstDt[5:7]

        yyyyPath = tabPath + "/" + rgstYYYY
        os.makedirs(yyyyPath , exist_ok= True)

        lastPath = yyyyPath + "/" + rgstMM + "월"
        os.makedirs(lastPath , exist_ok= True)
        
        fileDiv = accordian.find("div" , {"class": "accordion__contents"})
        setExcelValue(sheetPath , row , '구분' , "수시") #엑셀 셀 값 저장(수시)
        setExcelValue(sheetPath , row , '작성일' , rgstDt.replace("-","")) #엑셀 셀 값 저장(작성일)
        setExcelValue(sheetPath , row , '내용' , str(fileDiv)) #엑셀 셀 값 저장(내용)

        file = fileDiv.find("a", title = "다운로드")
        mainTopic = accordian.find("a" , {"class": "accordion__pointer"} )["title"]
        setExcelValue(sheetPath , row , '제목' , mainTopic) #엑셀 셀 값 저장(제목)

        if file != None :

            fileDownLoadUrl = file["href"].strip()        #다운로드할 파일 url
            subTopicP = fileDiv.find("p")
        
            subTopic = ''
            if subTopicP != None :
                if subTopicP.text.strip() != '':
                    subTopic = subTopicP.text.strip()
            
            saveName = rgstDt + "_" + mainTopic + ".pdf"   #기본 파일명은 메인주제를 따라 감

            if subTopic != '' :
                saveName = rgstDt + "_" + subTopic + ".pdf"  #디테일한 주제가 있으면 파일명으로 설정

            if fileDownLoadUrl.find(pruMainUrl) == -1 :
                fileDownLoadUrl = pruMainUrl + fileDownLoadUrl

            downloadPath = lastPath + "/" + saveName      #저장 경로

            try:
                download(fileDownLoadUrl , downloadPath)
                
                setExcelValue(sheetPath , row , '첨부파일' , downloadPath) #엑셀 셀 값 저장(첨부파일)
                print("success : " , downloadPath)
            except urllib.error.HTTPError as e:
                print("failed:", e)
        
        row += 1 #로우 증가
    dataTempltExcel.save('output/test.xlsx')  #엑셀 다른이름 저장 
   
    return

def cmpyInformationTable(tabInfo , tabPath , tabId) : #경영공시 테이블형식
    cmpyInformationTable = tabInfo.select('.table-holder table')[0]
    tables = cmpyInformationTable.select('tr')

    if tabId == 'regular' :
        sheetPath = dataTempltExcel.get_sheet_by_name("경영공시(정기,수시,결산,감사)")   #엑셀 시트명
            
    elif tabId == 'governance' :
        sheetPath = dataTempltExcel.get_sheet_by_name("경영공시(지배구조)")   #엑셀 시트명
    
    row = sheetPath.max_row + 1#엑셀 저장 로우 시작 (마지막 로우 조회) 

    for table in tables :
        dept1List = table.findAll("td", {"class": "va-t"}) #구분 년도(제목)
        
        if (dept1List) :   #경영공시의 정기/지배구조 공지와 같이 , 연도별 구분이 필요할 때 
            
            if tabId == 'regular' :
                yyyy = dept1List[0].text.strip().replace("년" , "")  #년도
                yyyyCell = yyyy
                setExcelValue(sheetPath , row , '구분' , "정기") #엑셀 셀 값 저장(정기)
            
            elif tabId == 'governance' :
                yyyy = dept1List[0].text.strip()[:4]  #년도
                yyyyCell = dept1List[0].text.strip().replace("-","")  #YYYYMMDD

            lastPath = tabPath + "/" + yyyy

            os.makedirs(lastPath , exist_ok= True)

            file = table.find("td", {"class": "ta-c"})
            
            setExcelValue(sheetPath , row , '제목' , dept1List[1].text.strip()) #엑셀 셀 값 저장(제목)
            setExcelValue(sheetPath , row , '작성일' , yyyyCell) #엑셀 셀 값 저장(작성일)

            if file.find("a") != None :
                if tabId == 'regular' :
                    saveName = dept1List[1].text.strip() #제목  #저장할 파일명
                
                elif tabId == 'governance' :
                    saveName = dept1List[0].text.strip() + "_" + dept1List[1].text.strip() #날짜_제목  #저장할 파일명

                fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url
                
                if fileDownLoadUrl.find(pruMainUrl) == -1 :
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                if saveName.find(".pdf") == -1 :
                    saveName += ".pdf"

                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath)
                    
                    setExcelValue(sheetPath , row , '첨부파일' , downloadPath) #엑셀 셀 값 저장(첨부파일 저장경로)

                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)
            
            row += 1#엑셀 로우 
    
    dataTempltExcel.save('output/test.xlsx')  #엑셀 다른이름 저장 


    return

def socialContribution(tabInfo , tabPath , tabId) : #사회공헌공시(공익법인 등 자산의 무상양도 공시)
    socialContribution = tabInfo.select('.table-holder table')[0]
    tables = socialContribution.select('tr')

    for table in tables :
        dept1 = table.find("td", {"class": "va-t"}) #구분 제목 폴더
        
        if dept1 != None :
            lastFolderNm = dept1.text.strip()
            fileAndNm = table.findAll("td", {"class": "ta-c"})
            file = fileAndNm[0]

            if tabId == 'disclosure' :  # 공시는 공시명(공시날짜)
                lastPath = tabPath + "/" + lastFolderNm
                os.makedirs(lastPath , exist_ok= True)

                #첨부파일 : fileAndNm[0] , 공시날짜 (첨부파일명) : fileAndNm[1] 
                fileNm = fileAndNm[1]
                saveName = fileNm.text.strip()  #저장할 파일명 (공시날짜)

            elif tabId == 'regulations' :  #규정은 제목을 파일명으로 저장
                lastPath = tabPath 
                saveName = lastFolderNm 

            if file.find("a") != None :

                fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url
                
                if fileDownLoadUrl.find(pruMainUrl) == -1 :
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                if tabId == 'disclosure' :  # 공시는 pdf
                    saveName += '.pdf'
                elif tabId == 'regulations' :  #규정은 docx
                    saveName += '.docx'

                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath)
                    
                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)

               
    return

def variableInsurance(tabInfo , tabPath) :  #변액보험공시 , 사회공헌공시 (사회공헌 관련 규정)
    variableInsurance = tabInfo.select('.table-holder table')[0]
    tables = variableInsurance.select('tr')

    for table in tables :
        dept1List = table.findAll("td", {"class": "va-t ta-l"}) #구분 1:1 상품명_판매기간
        
        if (dept1List) :   #변액보험공시 같이 , 상품명 / 판매기간 이 1:1 일 경우 
            
            lastFolderNm = ''
            for dept1 in dept1List :
                lastFolderNm = lastFolderNm + dept1.text.strip()
                
                if dept1List.index(dept1) != len(dept1List) - 1 :
                    lastFolderNm += '_'
            
            #글자수 체크 (63이하로 자르기)
            if len(lastFolderNm) > 63 :
                lastFolderNm = lastFolderNm[:63] + ".."
            
            lastPath = checkExistPathOrFile(tabPath + "/" + lastFolderNm.replace("/",","))  #동일한 파일명 있는지 확인 , '/'는 ','로 변환

            os.mkdir(lastPath)

            file = table.find("td", {"class": "ta-c"})

            if file.find("a") != None :
                saveName = file.find("a").find("img")["alt"]  #저장할 파일명

                fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url
                
                if fileDownLoadUrl.find(pruMainUrl) == -1 :
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                if saveName.find(".pdf") == -1 :
                    saveName += ".pdf"

                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath )
                    
                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)

               
        
    return


def product(tabInfo , optionPath):  #상품공시

    prod = tabInfo.select('.table-holder table')[0]
    tables = prod.select('tr')

    startFor = 1
    for table in tables :
        dept1 = table.find("td", {"class": "va-t"}) #구분
        
        if (dept1 != None) : 
            rowLen = int(dept1['rowspan'])
            
            dept1Strip = dept1.text.strip() #공백제거
            dept1Strip = checkExistPathOrFile(optionPath + "/" + dept1Strip)  #동일한 파일명 있는지 확인
            os.mkdir(dept1Strip)

            for i in range(startFor, startFor + rowLen):

                dept2 = tables[i].find("td", {"class": ""}) #상품명
                dept2Strip = dept2.text.strip() #공백제거

                dept2Url = dept2.find("a")["href"]
                routeUrl = pruMainUrl + dept2Url  #이동할 url

                saveFilePath = dept1Strip + "/" + dept2Strip 
                saveFilePath = checkExistPathOrFile(saveFilePath)  #동일한 파일명 있는지 확인
                os.makedirs(saveFilePath)
                clickDept2(routeUrl , saveFilePath)

            startFor += rowLen

    return 

def clickDept2(url , filePath) :
    soup = getPageSourceHtml(url)
    
    prod = soup.select('.table-holder table')[0]
    tables = prod.select('tr')

    startFor = 1
    for table in tables :
        dept1 = table.find("td", {"class": "va-t"}) #구분 1:n 상품명(판매기간)
        
        if (dept1 != None) :   #상품공시같이 , 상품명 / 판매기간 이 1:n 일 경우 
            rowLen = int(dept1['rowspan'])
            
            dept1Strip = dept1.text.strip() #공백제거

            dept3Path = checkExistPathOrFile(filePath + "/" + dept1Strip)  #동일한 파일명 있는지 확인
            os.mkdir(dept3Path)

            for i in range(startFor, startFor + rowLen):

                dept2 = tables[i].find("td", {"class": ""}) #상품명
                dept2Strip = dept2.text.strip()

                lastPath = checkExistPathOrFile(dept3Path + "/" + dept2Strip)  #동일한 파일명 있는지 확인
                os.makedirs(lastPath)

                fileList = tables[i].findAll("td", {"class": "ta-c"})

                for file in fileList : 
                    if file.find("a") != None :
                        saveName = file.find("a").find("div").find("img")["alt"]  #저장할 파일명

                        fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url

                        downloadPath = lastPath + "/" + saveName + ".pdf"       #저장 경로

                        try:
                            download(fileDownLoadUrl , downloadPath )
                            
                            print("success : " , downloadPath)
                        except urllib.error.HTTPError as e:
                            print("failed:", e)


            startFor += rowLen
        

    return 

def checkExistPathOrFile(pathOrFile) :  #파일명 존재여부 확인 후 있으면 '파일명(1)' 이런식으로 처리
    newPath = pathOrFile
    uniq=1

    while os.path.exists(newPath) :  #동일한 파일명 없을 때까지 반복
        newPath = '%s(%d)' % (pathOrFile,uniq)
        uniq += 1

    return newPath

def download(url, file_name = None):   #파일 다운로드 (다운로드할 파일 url , 저장경로(파일 이름))
            if not file_name:
                file_name = url.split('/')[-1]

            with open(file_name, "wb") as file:   
                    response = get(url)               
                    file.write(response.content) 

def getPageSourceHtml(url) :  # 페이지 소스 html변환
    # driver = webdriver.Chrome('./chromedriver')
    path = url
    chromeDriver.get(path)

    html = chromeDriver.page_source # html을 문자열로 가져온다.
    # response = get(url)
    # html = response.text
    # beautifulsoup 사용하기
    soup = BeautifulSoup(html,'html.parser')

    return soup

def setExcelValue(sheetPath , rowNum , cellNm , value):  #(시트 , 로우 , 칼럼명 , 데이터)
    cellNum = getCellTitleIndex(sheetPath[2] , cellNm)
    sheetPath.cell(rowNum,cellNum).value = value    #엑셀 셀 값 저장

def setExcelValueVI(sheetPath , rowNum , cellNm , value):  #(시트 , 로우 , 칼럼명 , 데이터)
    cellNum = getCellTitleIndex(sheetPath[3] , cellNm)
    sheetPath.cell(rowNum,cellNum).value = value    #엑셀 셀 값 저장

#템플릿 엑셀의 칼럼명의 인덱스 조회
def getCellTitleIndex(sheetRow , titleNm):
    
    def getValue(cell):
        return cell.value

    cellTitleList = list(map(getValue , sheetRow))
    index = cellTitleList.index(titleNm)

    if index == -1 :
        print("[fail] : do not find index")

    return index + 1



#손상된 파일 확인한 것
def checkFileYN() :
    excel = load_workbook('New Document 2022-09-14 144146.xlsx')#엑셀 템플릿
    sheetPath = excel.get_sheet_by_name("Sheet1")   #엑셀 시트명

    row = '1'
    while 100 != int(row) :
        cellNum = "A" + row

        try:
            if sheetPath[cellNum].value.strip() == "" :
                continue
        
            checkExistFile(sheetPath[cellNum].value.strip().replace("\\" , "/"))
            
        except AttributeError :
            print("except : " , row , " : " , sheetPath[cellNum].value)
        
        row = str(int(row) + 1)

    return

def checkExistFile(pathOrFile) :  

    if os.path.exists(pathOrFile) :  
        print(open(pathOrFile))
    else :
        print("fail : " , pathOrFile)
    

    return 


#주석 제외 후 실행
# selectTab('13343','https://www.prudential.co.kr/disclosure/variable-insurance-disclosure.aspx',['risk-indicator'])  #변액공시 (상품공시, 수시공시)['variable-insurance-product-disclosure', 'insurance-disclosure-at-any-time','risk-indicator']
# selectTab('13348','https://www.prudential.co.kr/disclosure/social-contribution-disclosure.aspx',['donation','social-service'])  #사회공헌공시 (기부 및 집행 세부내역 , 사회공헌 관련규정 , 공익법인 등 자산의 무상양도 공시) ['donation','social-service','regulations','disclosure']
# selectTab('13347','https://www.prudential.co.kr/disclosure/company-management-information.aspx',['regular' ,'governance', 'occasional'])   #경영공시 (정기/수시 경영공지 , 지배구조 공지) ['regular' ,'governance', 'occasional']
# selectTab('13342','https://www.prudential.co.kr/disclosure/product-disclosure.aspx',['currently-selling','discontinued'])   #상품공시 (판매상품 , 판매중지상품)
# selectTab('13339','https://www.prudential.co.kr/about-us/social-responsibility.aspx',['contribution-history'])   #회사소개 > 사회공헌 > 사회공헌활동 연혁
