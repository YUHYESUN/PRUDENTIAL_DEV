from tkinter.tix import Select
from bs4 import BeautifulSoup
import os
import urllib.request
from requests import get
from openpyxl import load_workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import Select
import time

'''
1. 공시실 최상위 메뉴 선택 (상품공시 , 경영공시 등등)
2. 탭 선택 (다운로드 필요한 탭id 값 하드코딩)
3. 검색창, 페이지 여부에 따라 생략될 수 있음
  3-1. 상품공시실) 주계약 , 특약 선택 (url에 파람 보냄 ~~?ct=01&tab=tabId)
  3-2. 경영공시실) 페이징 처리(페이지 번호 파람 보냄 ~~?page=10&tab=tabId)
4. 해당 탭 영역의 table 확인
5. 첨부파일 다운로드 
  5-1. 상품공시실) 상품공시 > 해당 탭명 > 주계약/특약 > 보험 구분명 > 보험 상품명 > 상세상품명 > 판매기간 > 파일 (지정된 파일명으로 저장) 
  5-2. 변액보험공시실) 변액보험공시 > 해당 탭명 > 상품명_판매기간 > 파일 (지정된 파일명으로 저장)
  5-3. 사회공헌공시실) 사회공헌공시 > 해당 탭명 > 
                         사회공헌관련 규정 ) 제목 > 파일 (제목으로 파일명 저장 확장자 : .docx)
                         자산무상양도 공시일 ) 제목 > 파일 (공시일자를 파일명으로 저장)
  5-4. 경영공시실) 경영공시 > 해당 탭명 > 
                       정기/지배구조 ) 연도별 > 파일 (정기공시는 제목을 , 지배구조공시는 날짜_제목을 파일명으로 저장)
                       수시경영공시  ) 연도별 > 월별 > 파일 (날짜_제목을 파일명으로 저장)
commit test123 123
'''

pruMainUrl = "https://www.prudential.co.kr"
pruMainUrl2 = "www.prudential.co.kr"
chromeDriver = webdriver.Chrome(ChromeDriverManager().install())
dataTempltExcel = load_workbook('dataTemplate.xlsx')#엑셀 템플릿
urlDict = {}

def selectTab(menuId , mainUrl , tabIdList) :

    url = mainUrl
    html = getPageSourceHtml(mainUrl) # html을 문자열로 가져온다.
            
    h2Nm = html.find("h2" , {"class" : "carousel__item-heading"})  #메뉴명 ex)상품공시 , 경영공시 ..
    
    #최상위폴더 생성
    h2NmStrip = checkExistPathOrFile("output" + "/" + h2Nm.text.strip())  #동일한 파일명 있는지 확인
    os.mkdir(h2NmStrip)

    if menuId == '13340': # 공지사항
        #페이지로 구성되어 있음 
        #페이지 1씩 더하다가 체크된 페이지랑 url에 입력된 페이지랑 맞지 않으면 스탑
        insertPage = "1"
        currentPage = html.find("strong" , {"class" : "SelectedPage"}).text

        while insertPage == currentPage :  #없는 페이지이면 스탑
            insertPage = str(int(insertPage) + 1)
            noticeAccordian(html, h2NmStrip)
            url = mainUrl + "?page=" + insertPage
            html = getPageSourceHtml(url)# html을 문자열로 가져온다.
            currentPage = html.find("strong" , {"class" : "SelectedPage"}).text

        return

    for tabId in tabIdList : 
        apiUrl = mainUrl + "?tab=" + tabId
        html = getPageSourceHtml(apiUrl) # html을 문자열로 가져온다.
        
        tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기
        tabNm = tabInfo.find("a" , {"class" : "accordion-tabs__item-toggle"}).find("span").text.strip()
        tabPath = checkExistPathOrFile(h2NmStrip + "/" + tabNm)  #동일한 파일명 있는지 확인
        # os.mkdir(tabPath)
        
        
        if menuId == '13341': # 퇴직연금공시
            if tabId == 'asset-liabilities': # 자산부채현황
                yearOptionId = 'ddlAssetLiabilitiesYear'
                monthOptionId = 'ddlAssetLiabilitiesMonth'
                year = '&aly='
                month = '&alm='
            else:
                yearOptionId = 'ddlAssetCompositionYear'
                monthOptionId = 'ddlAssetCompositionMonth'
                year = '&acy='
                month = '&acm='

            alyList = tabInfo.find("select", id = yearOptionId)
            yearOptions = alyList.findAll("option")

            yearSelect = Select(chromeDriver.find_element('id', yearOptionId))
            urlList = []
            for yearOption in yearOptions:
                yearOptionVal = yearOption["value"]
                yearSelect.select_by_value(yearOptionVal)
                time.sleep(1)
                selectMonth = Select(chromeDriver.find_element('id', monthOptionId))
                monthOptions = selectMonth.options

                for monthOption in monthOptions:
                    monthOptionVal = monthOption.text
                    url = apiUrl + year + yearOptionVal + month + monthOptionVal
                    urlList.append(url)
            
            retirementContribution(urlList, tabId) 

        elif menuId == "13342":   #상품공시

            if tabId == "currently-selling" :   #옵션아이디가 탭별로 다름
                optionId = "ddlContract"
            elif tabId == "discontinued" :
                optionId = "ddlDiscontinuedContract"

            optionSelect = tabInfo.find("select" , id = optionId)
            options = optionSelect.findAll("option")  #옵션 목록

            #주계약 (01), 특약 (02) url 2번 돌아야함
            for option in options : 
                optionVal = option["value"]
                
                optionPath = checkExistPathOrFile(tabPath + "/" + option.text)  #동일한 파일명 있는지 확인
                os.mkdir(optionPath)
                url = apiUrl + "&ct=" + optionVal

                html = getPageSourceHtml(url)# html을 문자열로 가져온다.

                tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기

                product(tabInfo , optionPath)

        elif menuId == "13343":  #변액공시
            
            variableInsurance(tabInfo , tabPath)
        
        elif menuId == '13347': #경영공시
            if tabId == 'regular' or tabId == 'governance' :
                cmpyInformationTable(tabInfo , tabPath , tabId)
            elif tabId == 'occasional' :
                #페이지로 구성되어 있음 
                #페이지 1씩 더하다가 체크된 페이지랑 url에 입력된 페이지랑 맞지 않으면 스탑
                insertPage = "1"
                currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text

                while insertPage == currentPage :  #없는 페이지이면 스탑
                    insertPage = str(int(insertPage) + 1)
                    cmpyInformationAccordian(tabInfo , tabPath )

                    url = apiUrl + "&page=" + insertPage
                    html = getPageSourceHtml(url)# html을 문자열로 가져온다.
                    tabInfo = html.find("div" , id = tabId) #탭에 해당된 테이블 찾기
                    currentPage = tabInfo.find("strong" , {"class" : "SelectedPage"}).text

        elif menuId == '13348': #사회공헌공시
            socialContribution(tabInfo , tabPath , tabId)
        
        
 

    return

def retirementContribution(urlList, tabId):
    if tabId == 'asset-liabilities': # 자산부채현황
        sheetPath = dataTempltExcel.get_sheet_by_name("자산부채현황(연금저축,자산연계,퇴직연금)")  #엑셀 시트명
        tabNm = '자산부채현황'
    else:
        sheetPath = dataTempltExcel.get_sheet_by_name("자산구성내역(연금저축,자산연계,퇴직연금)")   #엑셀 시트명
        tabNm = '자산구성내역'

    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 
    for url in urlList:
        html = getPageSourceHtml(url)# html을 문자열로 가져온다.
        
        year = url[url.find('&')+5:url.rfind('&')]
        month = url[url.rfind('=')+1:]
        if len(month) < 2:
            month = '0' + month

        if tabId == 'asset-liabilities': # 자산부채현황
            table = html.find("table", {"class":"--compact-2"})
        else:
            div = html.findAll("div", {"class":"--compact"})[3]
            table = div.find("table")
        
        sheetPath.cell(row,2).value = "퇴직연금"    # 공시구분
        sheetPath.cell(row,3).value = tabNm   # 현황구분
        sheetPath.cell(row,4).value = year   # 기준년
        sheetPath.cell(row,5).value = month   # 기준월
        sheetPath.cell(row,6).value = str(table)    # 내용 (CLOB)

        row += 1

    dataTempltExcel.save('output/retire.xlsx')  #엑셀 다른이름 저장 
    return
    
def noticeAccordian(info , path) : # 공지사항 아코디언 형식
    noticeAccordian = info.select('.section__pull-into-previous')[0]
    accordianList = noticeAccordian.select('.accordion')

    sheetPath = dataTempltExcel.get_sheet_by_name("구매입찰공시")   #엑셀 시트명
    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 
    
    for accordian in accordianList :
        rgstP = accordian.find("p" , {"class": "accordion__cover-small"} )
        rgstDt = rgstP.text.strip()
        rgstYYYY = rgstDt[:4]
        rgstMM = rgstDt[5:7]

        yyyyPath = path + "/" + rgstYYYY
        os.makedirs(yyyyPath , exist_ok= True)

        monthPath = yyyyPath + "/" + rgstMM + "월"
        os.makedirs(monthPath , exist_ok= True)

        lastPath = monthPath + "/" + rgstDt
        os.makedirs(lastPath , exist_ok= True)

        mainTopic = accordian.find("a" , {"class": "accordion__pointer"} )["title"] # 공지사항 제목
        fileDiv = accordian.find("div" , {"class": "accordion__contents"}) # 공지사항 내용
        sheetPath.cell(row,2).value = str(mainTopic)    # 제목
        sheetPath.cell(row,3).value = str(fileDiv)    # 내용  
        sheetPath.cell(row,4).value = rgstDt    # 등록일자  

        files = fileDiv.findAll("a")
        imgs = fileDiv.findAll("img")

        if len(files) > 0 or len(imgs) > 0:
            col = 5
            for file in files:
                fileDownLoadUrl = file["href"].strip()        # 다운로드할 파일 url
                if fileDownLoadUrl.find("getattachment") == -1: # 파일 아니면 넘어가 
                    continue

                docType = fileDownLoadUrl[fileDownLoadUrl.rfind("/"):]    # 문서형식 추출
                fileName = docType[:docType.find(".")]
                docType = docType[docType.find(".")+1:docType.rfind(".")]
                
                if fileDownLoadUrl.find(pruMainUrl2) == -1: # https://가 아닌 http:// 인것도 있어서 www부터 체크
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                saveName = fileName + "." + docType
                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath)
                    sheetPath.cell(row, col).value = downloadPath # 첨부파일경로
                    col += 1
                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)

            for file in imgs:
                fileDownLoadUrl = file["src"].strip()        # 다운로드할 파일 url
                docType = fileDownLoadUrl[fileDownLoadUrl.rfind("/"):]    # 문서형식 추출
                fileName = docType[:docType.find(".")]
                if fileDownLoadUrl.find("getattachment") != -1: # getattachment 형식이면
                    docType = docType[docType.find(".")+1:docType.rfind(".")]
                else:
                    docType = docType[docType.find(".")+1:]

                
                if fileDownLoadUrl.find(pruMainUrl2) == -1: # https://가 아닌 http:// 인것도 있어서 www부터 체크
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                saveName = fileName + "." + docType
                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath)
                    sheetPath.cell(row, col).value = downloadPath # 첨부파일경로
                    col += 1
                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)
        row += 1

    dataTempltExcel.save('output/notice.xlsx')  #엑셀 다른이름 저장 
    return

def cmpyInformationAccordian(tabInfo , tabPath) : #경영공시 아코디언 형식
    cmpyInformationAccordian = tabInfo.select('.panel__block')[0]
    accordianList = cmpyInformationAccordian.select('.accordion')
    
    for accordian in accordianList :
        rgstP = accordian.find("p" , {"class": "accordion__cover-small"} )
        rgstDt = rgstP.text.strip().replace("등록일 ","")
        rgstYYYY = rgstDt[:4]
        rgstMM = rgstDt[5:7]

        yyyyPath = tabPath + "/" + rgstYYYY
        os.makedirs(yyyyPath , exist_ok= True)

        lastPath = yyyyPath + "/" + rgstMM + "월"
        os.makedirs(lastPath , exist_ok= True)
        
        fileDiv = accordian.find("div" , {"class": "accordion__contents"})

        file = fileDiv.find("a", title = "다운로드")
        if file != None :

            fileDownLoadUrl = file["href"].strip()        #다운로드할 파일 url
            mainTopic = accordian.find("a" , {"class": "accordion__pointer"} )["title"]
            subTopicP = fileDiv.find("p")
        
            subTopic = ''
            if subTopicP != None :
                if subTopicP.text.strip() != '':
                    subTopic = subTopicP.text.strip()
            
            saveName = rgstDt + "_" + mainTopic + ".pdf"   #기본 파일명은 메인주제를 따라 감

            if subTopic != '' :
                saveName = rgstDt + "_" + subTopic + ".pdf"  #디테일한 주제가 있으면 파일명으로 설정

            if fileDownLoadUrl.find(pruMainUrl) == -1 :
                fileDownLoadUrl = pruMainUrl + fileDownLoadUrl

            downloadPath = lastPath + "/" + saveName      #저장 경로

            try:
                download(fileDownLoadUrl , downloadPath)
                
                print("success : " , downloadPath)
            except urllib.error.HTTPError as e:
                print("failed:", e)


    return

def cmpyInformationTable(tabInfo , tabPath , tabId) : #경영공시 테이블형식
    cmpyInformationTable = tabInfo.select('.table-holder table')[0]
    tables = cmpyInformationTable.select('tr')
    row = 4 #엑셀 저장 로우 시작 

    for table in tables :
        dept1List = table.findAll("td", {"class": "va-t"}) #구분 년도(제목)
        
        if (dept1List) :   #경영공시의 정기/지배구조 공지와 같이 , 연도별 구분이 필요할 때 
            
            if tabId == 'regular' :
                yyyy = dept1List[0].text.strip().replace("년" , "")  #년도
                yyyyCell = yyyy
            
            elif tabId == 'governance' :
                yyyy = dept1List[0].text.strip()[:4]  #년도
                yyyyCell = dept1List[0].text.strip().replace("-","")  #YYYYMMDD
                sheetPath = dataTempltExcel.get_sheet_by_name("경영공시(지배구조)")   #엑셀 시트명

            lastPath = tabPath + "/" + yyyy

            os.makedirs(lastPath , exist_ok= True)

            file = table.find("td", {"class": "ta-c"})
            
            #제목 : B , 작성일 : C
            sheetPath.cell(row,2).value = dept1List[1].text.strip()    #엑셀 셀 값 저장(제목)
            sheetPath.cell(row,3).value = yyyyCell    #엑셀 셀 값 저장(작성일)

            if file.find("a") != None :
                if tabId == 'regular' :
                    saveName = dept1List[1].text.strip() #제목  #저장할 파일명
                
                elif tabId == 'governance' :
                    saveName = dept1List[0].text.strip() + "_" + dept1List[1].text.strip() #날짜_제목  #저장할 파일명

                fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url
                
                if fileDownLoadUrl.find(pruMainUrl) == -1 :
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                if saveName.find(".pdf") == -1 :
                    saveName += ".pdf"

                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath)
                    #첨부파일 경로 (D)
                    sheetPath.cell(row,4).value = downloadPath    #엑셀 셀 값 저장

                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)
            
            row += 1#엑셀 로우 
    
    dataTempltExcel.save('output/test.xlsx')  #엑셀 다른이름 저장 


    return

def socialContribution(tabInfo , tabPath , tabId) : #사회공헌공시(공익법인 등 자산의 무상양도 공시)
    socialContribution = tabInfo.select('.table-holder table')[0]
    tables = socialContribution.select('tr')

    for table in tables :
        dept1 = table.find("td", {"class": "va-t"}) #구분 제목 폴더
        
        if dept1 != None :
            lastFolderNm = dept1.text.strip()
            fileAndNm = table.findAll("td", {"class": "ta-c"})
            file = fileAndNm[0]

            if tabId == 'disclosure' :  # 공시는 공시명(공시날짜)
                lastPath = tabPath + "/" + lastFolderNm
                os.makedirs(lastPath , exist_ok= True)

                #첨부파일 : fileAndNm[0] , 공시날짜 (첨부파일명) : fileAndNm[1] 
                fileNm = fileAndNm[1]
                saveName = fileNm.text.strip()  #저장할 파일명 (공시날짜)

            elif tabId == 'regulations' :  #규정은 제목을 파일명으로 저장
                lastPath = tabPath 
                saveName = lastFolderNm 

            if file.find("a") != None :

                fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url
                
                if fileDownLoadUrl.find(pruMainUrl) == -1 :
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                if tabId == 'disclosure' :  # 공시는 pdf
                    saveName += '.pdf'
                elif tabId == 'regulations' :  #규정은 docx
                    saveName += '.docx'

                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath)
                    
                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)

               
    return

def variableInsurance(tabInfo , tabPath) :  #변액보험공시 , 사회공헌공시 (사회공헌 관련 규정)
    variableInsurance = tabInfo.select('.table-holder table')[0]
    tables = variableInsurance.select('tr')

    for table in tables :
        dept1List = table.findAll("td", {"class": "va-t ta-l"}) #구분 1:1 상품명_판매기간
        
        if (dept1List) :   #변액보험공시 같이 , 상품명 / 판매기간 이 1:1 일 경우 
            
            lastFolderNm = ''
            for dept1 in dept1List :
                lastFolderNm = lastFolderNm + dept1.text.strip()
                
                if dept1List.index(dept1) != len(dept1List) - 1 :
                    lastFolderNm += '_'
            
            #글자수 체크 (63이하로 자르기)
            if len(lastFolderNm) > 63 :
                lastFolderNm = lastFolderNm[:63] + ".."
            
            lastPath = checkExistPathOrFile(tabPath + "/" + lastFolderNm.replace("/",","))  #동일한 파일명 있는지 확인 , '/'는 ','로 변환

            os.mkdir(lastPath)

            file = table.find("td", {"class": "ta-c"})

            if file.find("a") != None :
                saveName = file.find("a").find("img")["alt"]  #저장할 파일명

                fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url
                
                if fileDownLoadUrl.find(pruMainUrl) == -1 :
                    fileDownLoadUrl = pruMainUrl + fileDownLoadUrl
                
                if saveName.find(".pdf") == -1 :
                    saveName += ".pdf"

                downloadPath = lastPath + "/" + saveName      #저장 경로

                try:
                    download(fileDownLoadUrl , downloadPath )
                    
                    print("success : " , downloadPath)
                except urllib.error.HTTPError as e:
                    print("failed:", e)

               
        
    return


def product(tabInfo , optionPath):  #상품공시

    prod = tabInfo.select('.table-holder table')[0]
    tables = prod.select('tr')
    

    startFor = 1
    
    sheetPath = dataTempltExcel.get_sheet_by_name("상품공시")   #엑셀 시트명
    row = sheetPath.max_row + 1 #엑셀 로우 시작 (마지막 로우 조회) 
    for table in tables :
        dept1 = table.find("td", {"class": "va-t"}) #구분
        
        if (dept1 != None) : 
            rowLen = int(dept1['rowspan'])
            
            dept1Strip = dept1.text.strip() #공백제거
            dept1Strip = checkExistPathOrFile(optionPath + "/" + dept1Strip)  #동일한 파일명 있는지 확인
            os.mkdir(dept1Strip)

            for i in range(startFor, startFor + rowLen):

                dept2 = tables[i].find("td", {"class": ""}) #상품명
                dept2Strip = dept2.text.strip() #공백제거

                dept2Url = dept2.find("a")["href"]
                routeUrl = pruMainUrl + dept2Url  #이동할 url

                if routeUrl in urlDict:
                    continue
                urlDict[routeUrl] = 1

                saveFilePath = dept1Strip + "/" + dept2Strip 
                saveFilePath = checkExistPathOrFile(saveFilePath)  #동일한 파일명 있는지 확인
                os.makedirs(saveFilePath)
                
                row = clickDept2(routeUrl , saveFilePath, sheetPath, row)

            startFor += rowLen

    
    return 

def clickDept2(url , filePath, sheetPath, row) :
    

    soup = getPageSourceHtml(url)
    
    prod = soup.select('.table-holder table')[0]
    tables = prod.select('tr')

    
    startFor = 1
    for table in tables :
        dept1 = table.find("td", {"class": "va-t"}) #구분 1:n 상품명(판매기간)
        
        if (dept1 != None) :   #상품공시같이 , 상품명 / 판매기간 이 1:n 일 경우 
            rowLen = int(dept1['rowspan'])
            
            dept1Strip = dept1.text.strip() #공백제거

            dept3Path = checkExistPathOrFile(filePath + "/" + dept1Strip)  #동일한 파일명 있는지 확인
            os.mkdir(dept3Path)
            
            for i in range(startFor, startFor + rowLen):
                
                dept2 = tables[i].find("td", {"class": ""}) # 기간
                dept2Strip = dept2.text.strip()

                lastPath = checkExistPathOrFile(dept3Path + "/" + dept2Strip)  #동일한 파일명 있는지 확인
                os.makedirs(lastPath)
                
                duration = dept2Strip.split()
                if len(duration) < 3:
                    startDate = duration[0].replace("-","")
                    endDate = ""
                else:
                    startDate = duration[0].replace("-","")
                    endDate = duration[2].replace("-","")

                fileList = tables[i].findAll("td", {"class": "ta-c"})

                for file in fileList : 
                    if file.find("a") != None :
                        saveName = file.find("a").find("div").find("img")["alt"]  #저장할 파일명

                        fileDownLoadUrl = file.find("a")["href"]        #다운로드할 파일 url

                        downloadPath = lastPath + "/" + saveName + ".pdf"       #저장 경로

                        try:
                            download(fileDownLoadUrl , downloadPath )
                            sheetPath.cell(row,3).value = dept1Strip    # 상품명
                            sheetPath.cell(row,4).value = saveName[saveName.find("_")+1:saveName.rfind("_")]    # 문서구분
                            sheetPath.cell(row,5).value = startDate    # 시작일
                            sheetPath.cell(row,6).value = endDate    # 종료일
                            sheetPath.cell(row,7).value = downloadPath    #엑셀 셀 값 저장
                            row += 1#엑셀 로우
                            print("success : " , downloadPath)
                        except urllib.error.HTTPError as e:
                            print("failed:", e)
                

            startFor += rowLen
        
    dataTempltExcel.save('output/test.xlsx')  #엑셀 다른이름 저장 
    return row

def checkExistPathOrFile(pathOrFile) :  #파일명 존재여부 확인 후 있으면 '파일명(1)' 이런식으로 처리
    newPath = pathOrFile
    uniq=1

    while os.path.exists(newPath) :  #동일한 파일명 없을 때까지 반복
        newPath = '%s(%d)' % (pathOrFile,uniq)
        uniq += 1

    return newPath

def download(url, file_name = None):   #파일 다운로드 (다운로드할 파일 url , 저장경로(파일 이름))
            if not file_name:
                file_name = url.split('/')[-1]

            with open(file_name, "wb") as file:   
                    response = get(url)               
                    file.write(response.content) 

def getPageSourceHtml(url) :  # 페이지 소스 html변환
    # driver = webdriver.Chrome('./chromedriver')
    path = url
    chromeDriver.get(path)

    html = chromeDriver.page_source # html을 문자열로 가져온다.
    # response = get(url)
    # html = response.text
    # beautifulsoup 사용하기
    soup = BeautifulSoup(html,'html.parser')

    return soup

#주석 제외 후 실행
# selectTab('13343','https://www.prudential.co.kr/disclosure/variable-insurance-disclosure.aspx',['insurance-disclosure-at-any-time'])  #변액공시 (수시공시)
# selectTab('13348','https://www.prudential.co.kr/disclosure/social-contribution-disclosure.aspx',['regulations','disclosure'])  #사회공헌공시 (사회공헌 관련규정 , 공익법인 등 자산의 무상양도 공시)
# selectTab('13347','https://www.prudential.co.kr/disclosure/company-management-information.aspx',['governance'])   #경영공시 (정기/수시 경영공지 , 지배구조 공지) ['regular' ,'governance', 'occasional']
# selectTab('13342','https://www.prudential.co.kr/disclosure/product-disclosure.aspx',['currently-selling','discontinued'])   #상품공시 (판매상품 , 판매중지상품)
# selectTab('13340','https://www.prudential.co.kr/support-center/notices.aspx',[''])   # 공지사항
# selectTab('13341','https://www.prudential.co.kr/disclosure/retirement-disclosure.aspx',['asset-liabilities','asset-composition'])   # 퇴직연금(자산부채현황, 자산구성내역)
